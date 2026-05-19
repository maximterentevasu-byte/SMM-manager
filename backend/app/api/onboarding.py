import uuid
import io
import logging
from urllib.parse import quote

log = logging.getLogger(__name__)
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from pydantic import BaseModel
from typing import Optional
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from app.database import get_db
from app.models.models import User, Business
from app.api.auth import get_current_user
from app.services.onboarding_service import clarify_business_profile, parse_clarification_answer

router = APIRouter()

# ── Excel helpers ─────────────────────────────────────────────────────────────

PROFILE_FIELDS = [
    # (section, key, label, hint, is_list)
    ("Бизнес",     "name",                "Название бизнеса",                    "Официальное название компании",                                     False),
    ("Бизнес",     "niche",               "Ниша / сфера деятельности",           "Например: ресторан, интернет-магазин одежды, студия красоты",       False),
    ("Бизнес",     "usp",                 "УТП (уникальное торговое предложение)","Чем вы отличаетесь от конкурентов",                                 False),
    ("Бизнес",     "price_segment",       "Ценовой сегмент",                     "economy / middle / premium / luxury",                               False),
    ("Бизнес",     "geo",                 "Город / район",                       "Москва, Марьино",                                                   False),
    ("Бизнес",     "address",             "Адрес",                               "ул. Ленина, 12 (если есть физический адрес)",                       False),
    ("Бизнес",     "contact_info",        "Контакты",                            "Телефон, сайт, Telegram-бот через запятую",                         False),
    ("Бизнес",     "products",            "Товары и услуги",                     "Каждый пункт через ; (Пицца; Паста; Суп)",                          True),
    ("Акции",      "active_promotions",   "Акции и спецпредложения на месяц",    "Оставьте пустым если акций нет",                                    False),
    ("Аудитория",  "audience_primary",    "Основная целевая аудитория",          "Пол, возраст, интересы, образ жизни",                               False),
    ("Аудитория",  "audience_non_target", "Кто НЕ является вашей ЦА",           "Кого точно не нужно привлекать",                                    False),
    ("Аудитория",  "audience_pains",      "Боли клиентов / задачи продукта",     "Через ; (Нет времени готовить; Хочется вкусно и быстро)",           True),
    ("Аудитория",  "audience_objections", "Возражения клиентов",                 "Через ; (Дорого; Не знаю о вас; Не доверяю)",                       True),
    ("Конкуренты", "competitors",         "Конкуренты",                          "Название|Сайт через ; (МакДак|mcdonalds.ru; KFC|kfc.ru)",           True),
    ("Платформы",  "platforms",           "Платформы публикаций",                "Через ; (telegram; vk; ok)",                                        True),
    ("Бренд",      "brand_voice",         "Голос бренда / тональность",          "friendly / expert / innovator / family / luxury / fun",             False),
    ("Бренд",      "visual_style",        "Визуальный стиль",                    "Тёплые цвета, живые фото, минимализм...",                           False),
    ("Бренд",      "content_restrictions","Ограничения контента",                "Через ; (Политика; Алкоголь; Скидки)",                              True),
    ("Цели",       "business_goals",      "Цели бизнеса на 6 месяцев",           "Рост продаж на 30%, открыть вторую точку...",                       False),
    ("Цели",       "new_directions",      "Новые направления / продукты",        "Что планируете запустить в ближайшие месяцы",                       False),
    ("Цели",       "smm_metrics",         "Ключевые показатели SMM",             "Через ; (Подписчики; Охваты/Показы; ER вовлечённость ЦА)",          True),
]

_LABEL_TO_KEY: dict[str, str] = {row[2]: row[1] for row in PROFILE_FIELDS}
_KEY_IS_LIST:  dict[str, bool] = {row[1]: row[4] for row in PROFILE_FIELDS}


def _build_workbook(profile: dict | None) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Анкета бизнеса"

    hdr_fill  = PatternFill("solid", fgColor="1A1A1A")
    sec_fill  = PatternFill("solid", fgColor="F1EFE8")
    val_fill  = PatternFill("solid", fgColor="FFFEF7")
    hint_fill = PatternFill("solid", fgColor="FAFAF8")
    thin      = Side(style="thin", color="E0DED8")
    brd       = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 52

    ws.append(["Блок", "Поле", "Значение", "Подсказка / Пример"])
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = brd
    ws.row_dimensions[1].height = 24

    prev_section = None
    for section, key, label, hint, is_list in PROFILE_FIELDS:
        value = ""
        if profile:
            raw = profile.get(key)
            if is_list and isinstance(raw, list):
                if key == "competitors":
                    value = "; ".join(
                        f"{c.get('name','') if isinstance(c, dict) else c}|{c.get('url','') if isinstance(c, dict) else ''}"
                        for c in raw
                    )
                else:
                    value = "; ".join(str(x) for x in raw if x)
            elif raw is not None:
                value = str(raw)

        row_n = ws.max_row + 1
        ws.append([section if section != prev_section else "", label, value, hint])
        prev_section = section
        r = ws[row_n]

        r[0].fill = sec_fill;  r[0].font = Font(bold=True, color="555555", size=10)
        r[0].alignment = Alignment(horizontal="center", vertical="top", wrap_text=True); r[0].border = brd

        r[1].font = Font(bold=True, color="1A1A1A", size=11)
        r[1].alignment = Alignment(vertical="top", wrap_text=True); r[1].border = brd

        r[2].fill = val_fill; r[2].font = Font(color="1A1A1A", size=11)
        r[2].alignment = Alignment(vertical="top", wrap_text=True); r[2].border = brd

        r[3].fill = hint_fill; r[3].font = Font(color="999999", size=10, italic=True)
        r[3].alignment = Alignment(vertical="top", wrap_text=True); r[3].border = brd

        ws.row_dimensions[row_n].height = 28

    ws.freeze_panes = "A2"
    return wb


def _parse_workbook(wb: openpyxl.Workbook) -> dict:
    ws = wb.active
    result: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        label = str(row[1] or "").strip()
        raw   = row[2]
        value = str(raw).strip() if raw is not None else ""
        key   = _LABEL_TO_KEY.get(label)
        if not key or not value:
            continue
        if _KEY_IS_LIST.get(key):
            if key == "competitors":
                result[key] = [
                    {"name": p.split("|")[0].strip(), "url": p.split("|")[1].strip() if "|" in p else ""}
                    for p in value.split(";") if p.strip()
                ]
            else:
                result[key] = [x.strip() for x in value.split(";") if x.strip()]
        else:
            result[key] = value
    return result


class BusinessProfileRequest(BaseModel):
    name: str
    niche: str
    usp: str
    price_segment: str
    geo: str
    address: Optional[str] = None
    contact_info: Optional[str] = None
    products: list[str] = []
    active_promotions: Optional[str] = None
    audience_primary: str
    audience_non_target: Optional[str] = None
    audience_pains: list[str]
    audience_objections: list[str]
    competitors: list[dict]
    platforms: list[str]
    platform_goals: dict
    brand_voice: str
    brand_voice_examples: list[str]
    visual_style: str
    content_restrictions: list[str]
    logo_url: Optional[str] = None
    brand_colors: list[str] = []
    business_goals: Optional[str] = None
    new_directions: Optional[str] = None
    smm_metrics: list[str] = []
    brand_assets_labels: list[dict] = []


@router.post("/save-profile/{business_id}")
async def save_profile(
    business_id: str,
    profile_data: BusinessProfileRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    profile = profile_data.model_dump()
    business = None

    # Если business_id не "new" — ищем по конкретному ID
    if business_id != "new":
        try:
            result = await db.execute(
                select(Business).where(
                    Business.id == business_id,
                    Business.user_id == current_user.id
                )
            )
            business = result.scalar_one_or_none()
        except Exception:
            business = None

    # Если не нашли по ID (или "new") — ищем любой существующий бизнес пользователя.
    # Это предотвращает создание дублирующего бизнеса при потере localStorage.
    if business is None:
        try:
            result = await db.execute(
                select(Business).where(Business.user_id == current_user.id)
            )
            business = result.scalars().first()
        except Exception:
            business = None

    if business:
        business.profile = profile
        business.name = profile_data.name
    else:
        business = Business(
            id=uuid.uuid4(),
            user_id=current_user.id,
            name=profile_data.name,
            profile=profile
        )
        db.add(business)

    biz_id = str(business.id) if business.id else None
    try:
        await db.commit()
    except Exception as e:
        import traceback
        log.error("[save_profile] commit failed: %s\n%s", e, traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"DB error: {e}")
    return {"business_id": biz_id, "status": "saved"}


@router.post("/clarify/{business_id}")
async def get_clarifying_questions(
    business_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(Business).where(
            Business.id == business_id,
            Business.user_id == current_user.id
        )
    )
    business = result.scalar_one_or_none()
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")

    # Strip binary data from brand assets before sending to AI (avoids token limit)
    profile_for_ai = {**business.profile}
    if "brand_assets_labels" in profile_for_ai:
        profile_for_ai["brand_assets_labels"] = [
            {k: v for k, v in a.items() if k != "data"}
            for a in (profile_for_ai["brand_assets_labels"] or [])
        ]
    questions = await clarify_business_profile(profile_for_ai)
    return {"questions": questions}


@router.post("/quick-start")
async def quick_start(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Создаёт минимальный бизнес-профиль, если его ещё нет."""
    result = await db.execute(
        select(Business).where(Business.user_id == current_user.id)
    )
    existing = result.scalars().first()
    if existing:
        return {"business_id": str(existing.id), "status": "existing"}

    business = Business(
        id=uuid.uuid4(),
        user_id=current_user.id,
        name="",
        profile={"niche": "", "usp": "", "target_audience": "", "price_segment": "middle", "geo": ""},
    )
    db.add(business)
    await db.commit()
    return {"business_id": str(business.id), "status": "created"}


@router.post("/answer-clarification/{business_id}")
async def answer_clarification(
    business_id: str,
    body: dict,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(Business).where(
            Business.id == business_id,
            Business.user_id == current_user.id
        )
    )
    business = result.scalar_one_or_none()
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")

    profile_for_ai = {**business.profile}
    if "brand_assets_labels" in profile_for_ai:
        profile_for_ai["brand_assets_labels"] = [
            {k: v for k, v in a.items() if k != "data"}
            for a in (profile_for_ai["brand_assets_labels"] or [])
        ]
    updated_profile = await parse_clarification_answer(
        body["question"], body["answer"], profile_for_ai
    )
    business.profile = updated_profile
    await db.commit()
    return {"profile": updated_profile}


# ── Excel export / import ─────────────────────────────────────────────────────

@router.get("/export-template")
async def export_template():
    """Скачать пустой шаблон анкеты в формате Excel."""
    wb = _build_workbook(None)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=smm_profile_template.xlsx"},
    )


@router.get("/export-profile/{business_id}")
async def export_profile(
    business_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Выгрузить заполненную анкету бизнеса в Excel."""
    result = await db.execute(
        select(Business).where(Business.id == business_id, Business.user_id == current_user.id)
    )
    business = result.scalar_one_or_none()
    if not business:
        raise HTTPException(404, "Business not found")

    wb = _build_workbook(business.profile or {})
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = (business.name or "business").replace(" ", "_")[:40]
    encoded = quote(f"profile_{safe_name}.xlsx", safe="")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@router.post("/parse-excel")
async def parse_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Распарсить Excel-анкету и вернуть данные профиля (без сохранения)."""
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Принимаются только файлы .xlsx")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(400, "Не удалось прочитать файл. Убедитесь, что формат .xlsx")
    profile = _parse_workbook(wb)
    if not profile.get("name"):
        raise HTTPException(400, "В файле не найдено поле «Название бизнеса»")
    return {"profile": profile}