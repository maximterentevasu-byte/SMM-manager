import asyncio
import uuid
from io import BytesIO
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, func

from app.database import get_db
from app.models.models import (
    User, Business, PlatformConnection,
    TGWeeklyStats, VKWeeklyStats, ContentSlot, PlanStatus, Platform,
    TelegramPost, TelegramStory,
)
from app.api.auth import get_current_user
from app.services.publishers import decrypt_token, encrypt_token

router = APIRouter()

# Хранит api_id/api_hash между send-code → sign-in (только примитивы, не клиенты)
_pending_tg_meta: dict = {}


# ─── helpers ────────────────────────────────────────────────────────────────

async def _get_business(business_id: str, user: User, db: AsyncSession) -> Business:
    result = await db.execute(
        select(Business).where(Business.id == business_id, Business.user_id == user.id)
    )
    biz = result.scalar_one_or_none()
    if not biz:
        raise HTTPException(404, "Business not found")
    return biz


async def _get_connection(business_id: str, platform: str, db: AsyncSession) -> PlatformConnection | None:
    result = await db.execute(
        select(PlatformConnection).where(
            PlatformConnection.business_id == business_id,
            PlatformConnection.platform == platform,
            PlatformConnection.is_active == True,
        )
    )
    return result.scalar_one_or_none()


# ─── GET data ────────────────────────────────────────────────────────────────

@router.get("/{business_id}/tg")
async def get_tg_analytics(
    business_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_business(business_id, current_user, db)
    result = await db.execute(
        select(TGWeeklyStats)
        .where(TGWeeklyStats.business_id == business_id)
        .order_by(TGWeeklyStats.week_start.desc())
    )
    rows = result.scalars().all()
    return [
        {
            "id": str(r.id),
            "channel_id": r.channel_id,
            "channel_name": r.channel_name,
            "week_start": r.week_start.date().isoformat(),
            "week_end": r.week_end.date().isoformat(),
            "collected_at": r.collected_at.isoformat(),
            **r.stats,
        }
        for r in rows
    ]


@router.get("/{business_id}/vk")
async def get_vk_analytics(
    business_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_business(business_id, current_user, db)
    result = await db.execute(
        select(VKWeeklyStats)
        .where(VKWeeklyStats.business_id == business_id)
        .order_by(VKWeeklyStats.week_start.desc())
    )
    rows = result.scalars().all()
    return [
        {
            "id": str(r.id),
            "group_id": r.group_id,
            "group_name": r.group_name,
            "week_start": r.week_start.date().isoformat(),
            "week_end": r.week_end.date().isoformat(),
            "collected_at": r.collected_at.isoformat(),
            **r.stats,
        }
        for r in rows
    ]


# ─── TG credentials ──────────────────────────────────────────────────────────

class TGCredentialsIn(BaseModel):
    api_id: int
    api_hash: str
    session: str


@router.get("/{business_id}/tg-credentials")
async def get_tg_credentials_status(
    business_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_business(business_id, current_user, db)
    conn = await _get_connection(business_id, "telegram", db)
    if not conn:
        return {"configured": False, "has_connection": False}
    return {
        "configured": bool(conn.tg_api_id and conn.tg_session_encrypted),
        "has_connection": True,
        "channel_name": conn.page_name,
    }


@router.get("/{business_id}/vk-credentials")
async def get_vk_credentials_status(
    business_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_business(business_id, current_user, db)
    conn = await _get_connection(business_id, "vk", db)
    if not conn:
        return {"configured": False, "has_connection": False}
    return {
        "configured": bool(conn.vk_user_token_encrypted),
        "has_connection": True,
        "group_name": conn.page_name,
    }


class VKCredentialsIn(BaseModel):
    user_token: str


class TGPhoneIn(BaseModel):
    phone: str


class TGCodeIn(BaseModel):
    phone: str
    code: str
    phone_code_hash: str


def _tg_send_code_sync(api_id: int, api_hash: str, phone: str) -> tuple[str, str]:
    """Отправляет код, возвращает (phone_code_hash, session_str).
    Сессию нужно передать в sign_in — иначе Telegram отклонит хэш как невалидный."""
    import asyncio as _aio
    from telethon import TelegramClient
    from telethon.sessions import StringSession

    loop = _aio.new_event_loop()
    _aio.set_event_loop(loop)

    async def _do():
        client = TelegramClient(StringSession(), api_id, api_hash)
        await client.connect()
        sent = await client.send_code_request(phone)
        session_str = client.session.save()
        await client.disconnect()
        return sent.phone_code_hash, session_str

    return loop.run_until_complete(_do())


def _tg_sign_in_sync(api_id: int, api_hash: str, phone: str, code: str, phone_code_hash: str, session_str: str = "") -> str:
    """Выполняет sign_in, переиспользуя сессию от send_code (тот же auth_key)."""
    import asyncio as _aio
    from telethon import TelegramClient
    from telethon.sessions import StringSession

    loop = _aio.new_event_loop()
    _aio.set_event_loop(loop)

    async def _do():
        client = TelegramClient(StringSession(session_str or None), api_id, api_hash)
        await client.connect()
        await client.sign_in(phone, code, phone_code_hash=phone_code_hash)
        new_session = client.session.save()
        await client.disconnect()
        return new_session

    return loop.run_until_complete(_do())


@router.post("/{business_id}/tg-send-code")
async def tg_send_code(
    business_id: str,
    req: TGPhoneIn,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_business(business_id, current_user, db)
    conn = await _get_connection(business_id, "telegram", db)
    if not conn:
        raise HTTPException(400, "Telegram-канал не подключён. Сначала подключи в разделе Платформы.")
    from app.config import settings
    api_id = int(settings.TG_API_ID) if settings.TG_API_ID else 0
    api_hash = settings.TG_API_HASH or ""
    if not api_id or not api_hash:
        raise HTTPException(
            400,
            "На сервере не настроен TG_API_ID / TG_API_HASH. "
            "Добавь переменные окружения в Railway (получить на my.telegram.org → API development tools)."
        )
    try:
        phone_code_hash, temp_session = await asyncio.to_thread(
            _tg_send_code_sync, api_id, api_hash, req.phone
        )
        _pending_tg_meta[business_id] = {"api_id": api_id, "api_hash": api_hash, "session": temp_session}
        return {"phone_code_hash": phone_code_hash, "status": "code_sent"}
    except Exception as e:
        raise HTTPException(400, f"Ошибка отправки кода: {str(e)}")


@router.post("/{business_id}/tg-sign-in")
async def tg_sign_in(
    business_id: str,
    req: TGCodeIn,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_business(business_id, current_user, db)
    conn = await _get_connection(business_id, "telegram", db)
    if not conn:
        raise HTTPException(400, "Telegram-канал не подключён.")
    from app.config import settings
    api_id = int(settings.TG_API_ID) if settings.TG_API_ID else 0
    api_hash = settings.TG_API_HASH or ""
    if not api_id or not api_hash:
        raise HTTPException(400, "TG_API_ID / TG_API_HASH не настроены на сервере.")
    temp_session = (_pending_tg_meta.get(business_id) or {}).get("session", "")
    try:
        session_str = await asyncio.to_thread(
            _tg_sign_in_sync, api_id, api_hash, req.phone, req.code, req.phone_code_hash, temp_session
        )
        _pending_tg_meta.pop(business_id, None)
        conn.tg_api_id = str(api_id)
        conn.tg_api_hash = api_hash
        conn.tg_session_encrypted = encrypt_token(session_str)
        await db.commit()
        return {"status": "connected"}
    except Exception as e:
        raise HTTPException(400, f"Ошибка входа: {str(e)}")


@router.post("/{business_id}/vk-credentials")
async def save_vk_credentials(
    business_id: str,
    req: VKCredentialsIn,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_business(business_id, current_user, db)
    conn = await _get_connection(business_id, "vk", db)
    if not conn:
        raise HTTPException(400, "VK-сообщество не подключено. Сначала подключи в разделе Платформы.")
    conn.vk_user_token_encrypted = encrypt_token(req.user_token)
    await db.commit()
    return {"status": "saved"}


@router.post("/{business_id}/tg-credentials")
async def save_tg_credentials(
    business_id: str,
    req: TGCredentialsIn,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_business(business_id, current_user, db)
    conn = await _get_connection(business_id, "telegram", db)
    if not conn:
        raise HTTPException(400, "Telegram-канал не подключён. Сначала подключи в разделе Платформы.")
    from app.services.publishers import encrypt_token
    conn.tg_api_id = str(req.api_id)
    conn.tg_api_hash = req.api_hash
    conn.tg_session_encrypted = encrypt_token(req.session)
    await db.commit()
    return {"status": "saved"}


# ─── Collect: прямой вызов в рамках запроса через asyncio.to_thread ───────────

async def _save_vk(db: AsyncSession, business_id: str, vk: PlatformConnection) -> dict:
    """Собирает VK-аналитику и сохраняет в БД. Возвращает статус."""
    from app.services.analytics_vk import collect_vk_weekly

    community_token = decrypt_token(vk.token_encrypted)
    user_token = decrypt_token(vk.vk_user_token_encrypted) if vk.vk_user_token_encrypted else None

    if not user_token:
        return {
            "weeks": 0,
            "error": "Нужен VK user token для чтения стены. Добавь его в настройках аналитики.",
        }

    weekly, posts = await asyncio.to_thread(
        collect_vk_weekly, vk.external_page_id, community_token, user_token
    )
    if not weekly:
        return {"weeks": 0, "error": "Постов не найдено или VK API не вернул данных"}

    await db.execute(
        delete(VKWeeklyStats).where(
            VKWeeklyStats.business_id == business_id,
            VKWeeklyStats.group_id == vk.external_page_id.lstrip("-"),
        )
    )
    for w in weekly:
        ws_date = datetime.fromisoformat(w["week_start"])
        we_date = datetime.fromisoformat(w["week_end"])
        week_posts = [p for p in posts if w["week_start"] <= p["date"][:10] <= w["week_end"]]
        db.add(VKWeeklyStats(
            id=uuid.uuid4(),
            business_id=uuid.UUID(business_id),
            group_id=w["group_id"],
            group_name=w["group_name"],
            week_start=ws_date,
            week_end=we_date,
            stats={k: v for k, v in w.items()
                   if k not in ("group_id", "group_name", "week_start", "week_end")},
            posts=week_posts,
        ))
    await db.commit()
    return {"weeks": len(weekly)}


async def _save_tg_via_bot(db: AsyncSession, business_id: str, tg: PlatformConnection) -> dict:
    """
    Базовый сбор TG через бот-токен (без MTProto).
    Даёт: количество подписчиков + посты из ContentSlots за последние 12 недель.
    """
    import aiohttp
    from datetime import date, timedelta, datetime as dt

    bot_token = decrypt_token(tg.token_encrypted)

    async with aiohttp.ClientSession() as s:
        r1 = await s.get(
            f"https://api.telegram.org/bot{bot_token}/getChat",
            params={"chat_id": tg.external_page_id},
        )
        chat_resp = await r1.json()
        r2 = await s.get(
            f"https://api.telegram.org/bot{bot_token}/getChatMemberCount",
            params={"chat_id": tg.external_page_id},
        )
        members_resp = await r2.json()

    if not chat_resp.get("ok"):
        return {"weeks": 0, "error": chat_resp.get("description", "Bot API ошибка")}

    channel_name = chat_resp["result"].get("title", tg.page_name)
    subscribers = members_resp.get("result", 0) if members_resp.get("ok") else 0

    # Последние 12 завершённых недель из ContentSlots
    today = date.today()
    monday_this = today - timedelta(days=today.weekday())
    weeks_data = []
    for i in range(1, 13):
        ws = monday_this - timedelta(weeks=i)
        we = ws + timedelta(days=6)
        ws_dt = dt(ws.year, ws.month, ws.day, 0, 0, 0)
        we_dt = dt(we.year, we.month, we.day, 23, 59, 59)

        slots_res = await db.execute(
            select(ContentSlot).where(
                ContentSlot.business_id == business_id,
                ContentSlot.platform == Platform.telegram,
                ContentSlot.status == PlanStatus.published,
                ContentSlot.published_at >= ws_dt,
                ContentSlot.published_at <= we_dt,
            )
        )
        slots = slots_res.scalars().all()
        weeks_data.append({
            "week_start": ws.isoformat(),
            "week_end": we.isoformat(),
            "posts_count": len(slots),
        })

    # Всегда сохраняем хотя бы последнюю неделю (с subscriber count)
    save_weeks = [w for w in weeks_data if w["posts_count"] > 0] or [weeks_data[0]]

    await db.execute(
        delete(TGWeeklyStats).where(TGWeeklyStats.business_id == business_id)
    )
    for w in save_weeks:
        db.add(TGWeeklyStats(
            id=uuid.uuid4(),
            business_id=uuid.UUID(business_id),
            channel_id=tg.external_page_id,
            channel_name=channel_name,
            week_start=datetime.fromisoformat(w["week_start"]),
            week_end=datetime.fromisoformat(w["week_end"]),
            stats={
                "subscribers": subscribers,
                "posts_count": w["posts_count"],
                "total_views": None,
                "avg_views": None,
                "er_views_pct": None,
                "er_activity_pct": None,
                "quality_index": None,
                "_mode": "basic",
            },
            posts=[],
        ))
    await db.commit()
    return {"weeks": len(save_weeks), "mode": "basic", "subscribers": subscribers}


# Только RAW-данные от Telegram защищены от обнуления.
# Derived-метрики (ER, вираль) сюда НЕ входят — они всегда пересчитываются
# из актуальных (возможно защищённых) сырых данных.
_TG_PROTECT_RAW = {
    "subscribers",
    "total_views", "avg_views", "median_views",
    "avg_reactions", "avg_comments", "avg_reposts",
    "best_post_views", "worst_post_views",
}


def _recalc_derived(stats: dict) -> dict:
    """Пересчитывает производные метрики из (защищённых) сырых данных."""
    avg_views = stats.get("avg_views") or 0
    subscribers = stats.get("subscribers") or 0
    avg_reactions = stats.get("avg_reactions") or 0
    avg_comments = stats.get("avg_comments") or 0
    avg_reposts = stats.get("avg_reposts") or 0
    total_views = stats.get("total_views") or 0
    posts_count = stats.get("posts_count") or 1

    engagement_per_post = avg_reactions + avg_comments + avg_reposts
    stats["engagement_per_post"] = round(engagement_per_post, 1)

    stats["er_activity_pct"] = (
        round(engagement_per_post / avg_views * 100, 2) if avg_views > 0 else 0
    )

    # er_reach_pct только если есть subscriber-данные (только последняя неделя)
    if subscribers > 0 and avg_views > 0:
        stats["er_reach_pct"] = round(avg_views / subscribers * 100, 2)

    total_reposts = avg_reposts * posts_count
    stats["virality_pct"] = (
        round(total_reposts / total_views * 100, 3) if total_views > 0 else 0
    )

    return stats


async def _ai_week_status(week: dict, prev_week: dict | None) -> str:
    """Генерирует короткий AI-комментарий по неделе (Claude Haiku)."""
    try:
        from anthropic import AsyncAnthropic
        changes = []
        if prev_week:
            for key, label in [
                ("avg_views", "просмотры"), ("er_reach_pct", "ER просм."),
                ("subscribers", "подписчики"), ("er_activity_pct", "ER акт."),
            ]:
                curr = week.get(key) or 0
                prev = prev_week.get(key) or 0
                if prev > 0:
                    pct = (curr - prev) / prev * 100
                    direction = "▲" if pct > 0 else "▼"
                    changes.append(f"{label}: {direction}{abs(pct):.1f}%")

        context = (
            f"Неделя {week.get('week_start')}—{week.get('week_end')}\n"
            f"Посты: {week.get('posts_count', 0)}, Подписчики: {week.get('subscribers', '?')}\n"
            f"Ср. просмотр: {week.get('avg_views', 0):.0f}, "
            f"Медиана: {week.get('median_views', 0)}\n"
            f"ER просм.: {week.get('er_reach_pct', 0):.2f}%, "
            f"ER акт.: {week.get('er_activity_pct', 0):.2f}%\n"
            f"Лучший пост: {week.get('best_post_views', 0)} просм., "
            f"Худший: {week.get('worst_post_views', 0)} просм.\n"
            f"Виральность: {week.get('virality_pct', 0):.3f}%"
        )
        if changes:
            context += "\nК прошлой неделе: " + ", ".join(changes)

        client = AsyncAnthropic()
        msg = await client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=120,
            system=(
                "Ты SMM-аналитик. Дай 1-2 предложения — короткое операционное заключение по неделе "
                "на русском. Только факты, цифры, тренды. Без воды и рекомендаций."
            ),
            messages=[{"role": "user", "content": context}],
        )
        return msg.content[0].text.strip()
    except Exception:
        return ""


async def _upsert_tg_weekly(
    db: AsyncSession, business_id: str, channel_id: str,
    channel_name: str, weekly: list[dict], posts: list[dict],
) -> int:
    """Upsert недельной статистики: обновляет существующие недели, вставляет новые.
    Поля из _TG_PROTECT_ZERO не обнуляются если TG вернул 0."""
    biz_uuid = uuid.UUID(business_id)

    # Загружаем все существующие недели разом
    res = await db.execute(
        select(TGWeeklyStats).where(
            TGWeeklyStats.business_id == biz_uuid,
            TGWeeklyStats.channel_id == channel_id,
        )
    )
    existing_map: dict[str, TGWeeklyStats] = {
        row.week_start.date().isoformat(): row
        for row in res.scalars().all()
    }

    # Определяем недели, которым нужен AI-статус:
    # — новые (нет в БД) и последние 2 недели (данные свежие, статус мог устареть)
    sorted_weekly = sorted(weekly, key=lambda w: w["week_start"])
    recent_starts = {w["week_start"] for w in sorted_weekly[-2:]}
    needs_ai: list[tuple[int, dict, dict | None]] = []
    for i, w in enumerate(sorted_weekly):
        ws = w["week_start"]
        existing = existing_map.get(ws)
        existing_status = (existing.stats or {}).get("ai_status", "") if existing else ""
        if not existing_status or ws in recent_starts:
            prev = sorted_weekly[i - 1] if i > 0 else None
            needs_ai.append((i, w, prev))

    # Генерируем AI статусы параллельно (с семафором 5)
    sem = asyncio.Semaphore(5)

    async def _gen(w: dict, prev: dict | None) -> str:
        async with sem:
            return await _ai_week_status(w, prev)

    ai_results: list[str] = [""] * len(sorted_weekly)
    if needs_ai:
        statuses = await asyncio.gather(*[_gen(w, prev) for _, w, prev in needs_ai])
        for (idx, _, _), status in zip(needs_ai, statuses):
            ai_results[idx] = status

    # Upsert
    for i, w in enumerate(sorted_weekly):
        ws_date = datetime.fromisoformat(w["week_start"])
        we_date = datetime.fromisoformat(w["week_end"])
        week_posts = [p for p in posts if w["week_start"] <= p["date"][:10] <= w["week_end"]]

        new_stats = {k: v for k, v in w.items()
                     if k not in ("channel_id", "channel_name", "week_start", "week_end")}
        if ai_results[i]:
            new_stats["ai_status"] = ai_results[i]

        existing = existing_map.get(w["week_start"])
        if existing:
            old_stats = dict(existing.stats or {})
            # Защита: не обнуляем RAW-данные от Telegram (подписчики, просмотры и т.д.)
            for k in _TG_PROTECT_RAW:
                new_val = new_stats.get(k)
                old_val = old_stats.get(k)
                if old_val and not new_val:
                    new_stats[k] = old_val
            # Сохраняем старый ai_status если новый пустой
            if not new_stats.get("ai_status") and old_stats.get("ai_status"):
                new_stats["ai_status"] = old_stats["ai_status"]
            # Всегда пересчитываем производные метрики из (защищённых) сырых данных
            _recalc_derived(new_stats)
            existing.stats = new_stats
            existing.channel_name = channel_name
            existing.collected_at = datetime.utcnow()
            if week_posts:
                existing.posts = week_posts
        else:
            _recalc_derived(new_stats)
            db.add(TGWeeklyStats(
                id=uuid.uuid4(),
                business_id=biz_uuid,
                channel_id=channel_id,
                channel_name=channel_name,
                week_start=ws_date,
                week_end=we_date,
                stats=new_stats,
                posts=week_posts,
            ))

    await db.commit()
    return len(weekly)


async def _save_tg(db: AsyncSession, business_id: str, tg: PlatformConnection) -> dict:
    """Собирает TG-аналитику. MTProto если настроен, иначе базовый Bot API режим."""
    from app.config import settings

    api_id = int(tg.tg_api_id) if tg.tg_api_id else settings.TG_API_ID
    api_hash = tg.tg_api_hash if tg.tg_api_hash else settings.TG_API_HASH
    session = (
        decrypt_token(tg.tg_session_encrypted)
        if tg.tg_session_encrypted
        else settings.TG_STRING_SESSION
    )

    if api_id and session:
        from app.services.analytics_tg import collect_tg_weekly, _TG_AUTH_KEYWORDS
        try:
            weekly, posts = await asyncio.to_thread(
                collect_tg_weekly, api_id, api_hash, session, tg.external_page_id
            )
        except Exception as e:
            err_lower = str(e).lower()
            if any(k in err_lower for k in _TG_AUTH_KEYWORDS):
                tg.tg_session_encrypted = None
                await db.commit()
                return {
                    "weeks": 0,
                    "error": f"Сессия Telegram устарела — переподключите аккаунт в настройках аналитики. ({e})",
                }
            if any(k in err_lower for k in ("eof", "connection", "reset", "broken pipe", "timed out", "incomplete")):
                fallback = await _save_tg_via_bot(db, business_id, tg)
                fallback["warning"] = f"MTProto временно недоступен, данные собраны через Bot API. ({e})"
                return fallback
            return {"weeks": 0, "error": str(e)}

        if not weekly:
            return {"weeks": 0, "error": "MTProto: постов не найдено"}

        count = await _upsert_tg_weekly(
            db, business_id, tg.external_page_id,
            weekly[0].get("channel_name", tg.page_name),
            weekly, posts,
        )
        return {"weeks": count, "mode": "full"}

    return await _save_tg_via_bot(db, business_id, tg)


@router.post("/{business_id}/collect")
async def collect_analytics(
    business_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_business(business_id, current_user, db)
    result: dict = {}

    vk = await _get_connection(business_id, "vk", db)
    if vk:
        try:
            result["vk"] = await _save_vk(db, business_id, vk)
        except Exception as e:
            result["vk"] = {"weeks": 0, "error": str(e)}

    tg = await _get_connection(business_id, "telegram", db)
    if tg:
        try:
            result["tg"] = await _save_tg(db, business_id, tg)
        except Exception as e:
            result["tg"] = {"weeks": 0, "error": str(e)}

    if not vk and not tg:
        return {"status": "no_connections", "detail": "Нет подключённых платформ"}

    return {"status": "done", **result}


# ─── Excel export ────────────────────────────────────────────────────────────

def _make_excel_tg(rows: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "TG_Недели"

    HEADERS = [
        "Неделя", "Канал", "Подписчики", "Постов",
        "Всего просмотров", "Ср. охват", "Медиана охвата",
        "Ср. реакции", "Ср. комм.", "Ср. репосты",
        "ER просмотры %", "ER активность %",
        "Вирусность %", "Вовлечённость /1000", "Индекс качества",
        "Лучший день", "Лучший час",
    ]

    hdr_fill = PatternFill("solid", fgColor="1A1A1A")
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 2):
        fill = PatternFill("solid", fgColor="F8F7F4" if ri % 2 == 0 else "FFFFFF")
        vals = [
            f"{row.get('week_start','')} — {row.get('week_end','')}",
            row.get("channel_name", ""),
            row.get("subscribers", ""),
            row.get("posts_count", ""),
            row.get("total_views", ""),
            row.get("avg_views", ""),
            row.get("median_views", ""),
            row.get("avg_reactions", ""),
            row.get("avg_comments", ""),
            row.get("avg_reposts", ""),
            row.get("er_views_pct", ""),
            row.get("er_activity_pct", ""),
            row.get("virality_pct", ""),
            row.get("engagement_per_1000", ""),
            row.get("quality_index", ""),
            row.get("best_day", ""),
            row.get("best_hour", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(ri, col, val)
            c.fill = fill

    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    ws2 = wb.create_sheet("TG_Посты")
    P_HEADERS = ["Канал", "Дата (EKB)", "ID", "Просмотры", "Реакции", "Комм.", "Репосты", "Текст"]
    for col, h in enumerate(P_HEADERS, 1):
        c = ws2.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr_fill
    for row in rows:
        for post in (row.get("posts") or []):
            ws2.append([
                row.get("channel_name", ""),
                post.get("date", ""),
                post.get("id", ""),
                post.get("views", 0),
                post.get("reactions", 0),
                post.get("comments", 0),
                post.get("reposts", 0),
                post.get("text", ""),
            ])
    ws2.column_dimensions["H"].width = 60

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_excel_vk(rows: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "VK_Недели"

    HEADERS = [
        "Неделя", "Сообщество", "Участников", "Постов",
        "Всего просмотров", "Ср. охват", "Медиана охвата",
        "Ср. лайки", "Ср. комм.", "Ср. репосты",
        "ER подписчики %", "ER просмотры %",
        "Вирусность %", "Индекс вовлечённости",
        "Прирост подписчиков", "Лучший день", "Лучший час",
    ]

    hdr_fill = PatternFill("solid", fgColor="1A1A1A")
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 2):
        fill = PatternFill("solid", fgColor="F8F7F4" if ri % 2 == 0 else "FFFFFF")
        vals = [
            f"{row.get('week_start','')} — {row.get('week_end','')}",
            row.get("group_name", ""),
            row.get("members", ""),
            row.get("posts_count", ""),
            row.get("total_views", ""),
            row.get("avg_views", ""),
            row.get("median_views", ""),
            row.get("avg_likes", ""),
            row.get("avg_comments", ""),
            row.get("avg_reposts", ""),
            row.get("er_subscribers_pct", ""),
            row.get("er_views_pct", ""),
            row.get("virality_pct", ""),
            row.get("engagement_index", ""),
            row.get("net_growth", "н/д"),
            row.get("best_day", ""),
            row.get("best_hour", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(ri, col, val)
            c.fill = fill

    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    ws2 = wb.create_sheet("VK_Посты")
    P_HEADERS = ["Сообщество", "Дата (EKB)", "ID", "Просмотры", "Лайки", "Комм.", "Репосты", "Текст"]
    for col, h in enumerate(P_HEADERS, 1):
        c = ws2.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr_fill
    for row in rows:
        for post in (row.get("posts") or []):
            ws2.append([
                row.get("group_name", ""),
                post.get("date", ""),
                post.get("id", ""),
                post.get("views", 0),
                post.get("likes", 0),
                post.get("comments", 0),
                post.get("reposts", 0),
                post.get("text", ""),
            ])
    ws2.column_dimensions["H"].width = 60

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/{business_id}/tg/export")
async def export_tg_excel(
    business_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_business(business_id, current_user, db)
    result = await db.execute(
        select(TGWeeklyStats)
        .where(TGWeeklyStats.business_id == business_id)
        .order_by(TGWeeklyStats.week_start.asc())
    )
    rows = result.scalars().all()
    if not rows:
        raise HTTPException(404, "Нет данных для экспорта")

    data = [{**r.stats, "posts": r.posts, "channel_name": r.channel_name,
             "week_start": r.week_start.date().isoformat(),
             "week_end": r.week_end.date().isoformat()} for r in rows]

    file_bytes = _make_excel_tg(data)
    return StreamingResponse(
        BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tg_analytics.xlsx"},
    )


@router.get("/{business_id}/vk/export")
async def export_vk_excel(
    business_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_business(business_id, current_user, db)
    result = await db.execute(
        select(VKWeeklyStats)
        .where(VKWeeklyStats.business_id == business_id)
        .order_by(VKWeeklyStats.week_start.asc())
    )
    rows = result.scalars().all()
    if not rows:
        raise HTTPException(404, "Нет данных для экспорта")

    data = [{**r.stats, "posts": r.posts, "group_name": r.group_name,
             "week_start": r.week_start.date().isoformat(),
             "week_end": r.week_end.date().isoformat()} for r in rows]

    file_bytes = _make_excel_vk(data)
    return StreamingResponse(
        BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vk_analytics.xlsx"},
    )


# ─── TG POSTS (per-post analytics) ───────────────────────────────────────────

def _collect_posts_sync(api_id: int, api_hash: str, session_str: str,
                        channel: str, max_id: int = 0) -> list[dict]:
    """Собирает посты из Telegram-канала. max_id=0 → последние 500; max_id>0 → старше max_id."""
    import asyncio as _aio
    from telethon import TelegramClient
    from telethon.sessions import StringSession
    from app.services.analytics_tg import _parse_channel_id

    loop = _aio.new_event_loop()
    _aio.set_event_loop(loop)

    async def _do():
        client = TelegramClient(StringSession(session_str or None), api_id, api_hash)
        await client.connect()
        posts = []
        try:
            # Нормализуем ID так же как _fetch_channel в analytics_tg.py
            numeric_id = _parse_channel_id(channel)
            full_id = int(f"-100{numeric_id}")

            try:
                entity = await client.get_entity(full_id)
            except Exception:
                await client.get_dialogs(limit=50)
                entity = await client.get_entity(full_id)
            subscribers = getattr(entity, "participants_count", 0) or 0
            ch_name = getattr(entity, "username", "") or getattr(entity, "title", "") or ""

            kwargs: dict = {"limit": 500}
            if max_id > 0:
                kwargs["max_id"] = max_id  # уходим глубже в историю

            async for msg in client.iter_messages(entity, **kwargs):
                if not msg or getattr(msg, "service", False):
                    continue

                views = msg.views or 0
                reacts = 0
                if msg.reactions:
                    reacts = sum(r.count for r in msg.reactions.results)
                reposts = msg.forwards or 0
                comments = (msg.replies.replies if msg.replies else 0)

                er = ((reacts + comments + reposts) / views * 100) if views > 0 else 0.0
                viral = (reposts / views * 100) if views > 0 else 0.0

                text = msg.text or ""
                mtype = "none"
                if msg.photo:
                    mtype = "photo"
                elif getattr(msg, "video", None) or getattr(msg, "video_note", None):
                    mtype = "video"
                elif getattr(msg, "voice", None):
                    mtype = "voice"
                elif msg.document:
                    mtype = "document"

                posts.append({
                    "post_id": msg.id,
                    "published_at": msg.date.replace(tzinfo=None),
                    "text": text,
                    "views": views,
                    "reactions": reacts,
                    "comments": comments,
                    "reposts": reposts,
                    "has_media": mtype != "none",
                    "media_type": mtype,
                    "er_pct": round(er, 2),
                    "virality_pct": round(viral, 4),
                    "has_question": "?" in text,
                    "subscribers": subscribers,
                    "channel_name": ch_name,
                })
        finally:
            await client.disconnect()
        return posts

    return loop.run_until_complete(_do())


async def _upsert_posts(business_id: str, posts: list[dict], db: AsyncSession) -> int:
    """Upsert TelegramPost rows — обновляет метрики если пост уже есть."""
    if not posts:
        return 0
    biz_uuid = uuid.UUID(str(business_id))
    for pd in posts:
        res = await db.execute(
            select(TelegramPost).where(
                TelegramPost.business_id == biz_uuid,
                TelegramPost.post_id == pd["post_id"],
            )
        )
        existing = res.scalar_one_or_none()
        if existing:
            for k, v in pd.items():
                setattr(existing, k, v)
            existing.updated_at = datetime.utcnow()
        else:
            db.add(TelegramPost(id=uuid.uuid4(), business_id=biz_uuid, **pd))
    await db.commit()
    return len(posts)


@router.get("/{business_id}/tg/posts")
async def get_tg_posts(
    business_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Возвращает все сохранённые посты (из БД, не из Telegram)."""
    await _get_business(business_id, current_user, db)
    result = await db.execute(
        select(TelegramPost)
        .where(TelegramPost.business_id == business_id)
        .order_by(TelegramPost.published_at.desc())
    )
    rows = result.scalars().all()
    return [
        {
            "post_id": r.post_id,
            "channel_name": r.channel_name,
            "published_at": r.published_at.isoformat(),
            "text": r.text,
            "views": r.views,
            "reactions": r.reactions,
            "comments": r.comments,
            "reposts": r.reposts,
            "has_media": r.has_media,
            "media_type": r.media_type,
            "er_pct": r.er_pct,
            "virality_pct": r.virality_pct,
            "has_question": r.has_question,
            "subscribers": r.subscribers,
            "updated_at": r.updated_at.isoformat(),
        }
        for r in rows
    ]


@router.post("/{business_id}/tg/posts/collect")
async def collect_tg_posts(
    business_id: str,
    deeper: bool = Query(False, description="True — уходим глубже в историю"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Собирает/обновляет посты из Telegram. deeper=True → продолжает с самого старого."""
    await _get_business(business_id, current_user, db)
    conn = await _get_connection(business_id, "telegram", db)
    if not conn or not conn.tg_session_encrypted:
        raise HTTPException(400, "Telegram не авторизован. Войдите через вкладку Аналитика → Telegram.")

    from app.config import settings
    api_id = int(settings.TG_API_ID) if settings.TG_API_ID else 0
    api_hash = settings.TG_API_HASH or ""
    if not api_id:
        raise HTTPException(400, "TG_API_ID / TG_API_HASH не настроены.")

    session_str = decrypt_token(conn.tg_session_encrypted)
    channel = conn.external_page_id or ""
    if not channel:
        raise HTTPException(400, "Идентификатор канала не найден. Переподключите Telegram в разделе Платформы.")

    # Определяем max_id: для deeper берём минимальный post_id из уже сохранённых
    max_id = 0
    if deeper:
        res = await db.execute(
            select(func.min(TelegramPost.post_id)).where(TelegramPost.business_id == business_id)
        )
        min_stored = res.scalar()
        if min_stored:
            max_id = min_stored - 1  # уходим глубже (older)

    try:
        posts = await asyncio.to_thread(
            _collect_posts_sync, api_id, api_hash, session_str, channel, max_id
        )
        count = await _upsert_posts(business_id, posts, db)
        return {"status": "ok", "collected": count, "deeper": deeper}
    except Exception as e:
        raise HTTPException(400, f"Ошибка сбора постов: {str(e)}")


# ─── TG STORIES ───────────────────────────────────────────────────────────────

def _collect_stories_sync(api_id: int, api_hash: str, session_str: str,
                          channel: str, offset_id: int = 0) -> list[dict]:
    """Собирает истории из Telegram-канала через архив Stories."""
    import asyncio as _aio
    from telethon import TelegramClient
    from telethon.sessions import StringSession
    from app.services.analytics_tg import _parse_channel_id

    loop = _aio.new_event_loop()
    _aio.set_event_loop(loop)

    async def _do():
        client = TelegramClient(StringSession(session_str or None), api_id, api_hash)
        await client.connect()
        stories_out = []
        try:
            numeric_id = _parse_channel_id(channel)
            full_id = int(f"-100{numeric_id}")

            try:
                entity = await client.get_entity(full_id)
            except Exception:
                await client.get_dialogs(limit=50)
                entity = await client.get_entity(full_id)

            ch_name = getattr(entity, "username", "") or getattr(entity, "title", "") or ""

            from telethon.tl.functions.stories import GetStoriesArchiveRequest
            result = await client(GetStoriesArchiveRequest(
                peer=entity,
                offset_id=offset_id,
                limit=100,
            ))

            for story in (result.stories or []):
                views_count = 0
                reactions_count = 0
                forwards_count = 0
                sv = getattr(story, "views", None)
                if sv:
                    views_count = getattr(sv, "views_count", 0) or 0
                    reactions_count = getattr(sv, "reactions_count", 0) or 0
                    forwards_count = getattr(sv, "forwards_count", 0) or 0

                has_media = False
                mtype = "none"
                media = getattr(story, "media", None)
                if media:
                    has_media = True
                    if hasattr(media, "photo") and media.photo:
                        mtype = "photo"
                    elif hasattr(media, "document") and media.document:
                        mtype = "video"

                er = ((reactions_count + forwards_count) / views_count * 100) if views_count > 0 else 0.0

                published = story.date
                if hasattr(published, "replace"):
                    published = published.replace(tzinfo=None)

                expires_raw = getattr(story, "expire_date", None)
                expires = None
                if isinstance(expires_raw, int):
                    from datetime import datetime as _dt
                    expires = _dt.utcfromtimestamp(expires_raw)
                elif expires_raw:
                    expires = expires_raw.replace(tzinfo=None) if hasattr(expires_raw, "replace") else None

                caption = getattr(story, "caption", "") or ""

                stories_out.append({
                    "story_id": story.id,
                    "channel_name": ch_name,
                    "published_at": published,
                    "expires_at": expires,
                    "caption": caption,
                    "views": views_count,
                    "reactions": reactions_count,
                    "forwards": forwards_count,
                    "has_media": has_media,
                    "media_type": mtype,
                    "er_pct": round(er, 2),
                })
        finally:
            await client.disconnect()
        return stories_out

    return loop.run_until_complete(_do())


async def _upsert_stories(business_id: str, stories: list[dict], db: AsyncSession) -> int:
    if not stories:
        return 0
    biz_uuid = uuid.UUID(str(business_id))
    # Метрики, которые могут обнуляться в Telegram после истечения истории —
    # перезаписываем только если новое значение больше уже сохранённого.
    _PRESERVABLE = ("views", "reactions", "forwards")
    for sd in stories:
        res = await db.execute(
            select(TelegramStory).where(
                TelegramStory.business_id == biz_uuid,
                TelegramStory.story_id == sd["story_id"],
            )
        )
        existing = res.scalar_one_or_none()
        if existing:
            for k, v in sd.items():
                if k in _PRESERVABLE:
                    # Сохраняем максимум: не затираем ненулевые данные нулями
                    old = getattr(existing, k, 0) or 0
                    setattr(existing, k, max(old, v or 0))
                else:
                    setattr(existing, k, v)
            # Пересчитываем ER на основе актуальных (preserved) значений
            v_views = existing.views or 0
            if v_views > 0:
                existing.er_pct = round((existing.reactions + existing.forwards) / v_views * 100, 2)
            existing.updated_at = datetime.utcnow()
        else:
            db.add(TelegramStory(id=uuid.uuid4(), business_id=biz_uuid, **sd))
    await db.commit()
    return len(stories)


@router.get("/{business_id}/tg/stories")
async def get_tg_stories(
    business_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_business(business_id, current_user, db)
    result = await db.execute(
        select(TelegramStory)
        .where(TelegramStory.business_id == business_id)
        .order_by(TelegramStory.published_at.desc())
    )
    rows = result.scalars().all()
    return [
        {
            "story_id": r.story_id,
            "channel_name": r.channel_name,
            "published_at": r.published_at.isoformat(),
            "expires_at": r.expires_at.isoformat() if r.expires_at else None,
            "caption": r.caption,
            "views": r.views,
            "reactions": r.reactions,
            "forwards": r.forwards,
            "has_media": r.has_media,
            "media_type": r.media_type,
            "er_pct": r.er_pct,
            "updated_at": r.updated_at.isoformat(),
        }
        for r in rows
    ]


def _rate_vk_posts_for_analysis(posts: list[dict], avg_views: float, avg_er: float) -> list[dict]:
    """Рейтинг VK-постов с поправкой на возраст."""
    from datetime import datetime as _dt
    now = _dt.utcnow()
    rated = []
    for p in posts:
        date_str = str(p.get("date", ""))
        try:
            pub = _dt.fromisoformat(date_str[:19].replace("T", " ").replace("Z", ""))
        except Exception:
            pub = now
        age_days = max((now - pub).days, 0)

        views    = p.get("views", 0) or 0
        likes    = p.get("likes", 0) or 0
        comments = p.get("comments", 0) or 0
        reposts  = p.get("reposts", 0) or 0
        er       = (likes + comments + reposts) / views * 100 if views > 0 else 0
        viral    = reposts / views * 100 if views > 0 else 0

        if age_days < 1:
            age_factor, age_label = 0.08, "опубликован сегодня — данных слишком мало"
        elif age_days < 3:
            age_factor, age_label = 0.30, f"{age_days} дн. назад — ещё набирает просмотры"
        elif age_days < 7:
            age_factor, age_label = 0.65, f"{age_days} дн. назад — статистика неполная"
        elif age_days < 14:
            age_factor, age_label = 0.90, None
        else:
            age_factor, age_label = 1.0, None

        adj_views   = views / age_factor if age_factor > 0 else views
        views_score = min((adj_views / max(avg_views, 1)) * 50, 100)
        er_score    = min((er / max(avg_er, 0.01)) * 30, 60)
        viral_score = min(viral * 5, 20)
        score = round(views_score + er_score + viral_score, 1)

        rated.append({**p, "score": score, "age_days": age_days,
                      "age_label": age_label, "adj_views": round(adj_views),
                      "er_pct": round(er, 2)})

    return sorted(rated, key=lambda x: x["score"], reverse=True)


@router.post("/{business_id}/vk/ai-analysis")
async def vk_ai_analysis(
    business_id: str,
    weeks: int = Query(8, description="0 = всё время"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Полный AI-разбор VK-сообщества: динамика, рейтинг постов, рекомендации."""
    await _get_business(business_id, current_user, db)

    from datetime import date, timedelta

    wq = select(VKWeeklyStats).where(VKWeeklyStats.business_id == business_id)
    if weeks > 0:
        cutoff = date.today() - timedelta(weeks=weeks)
        wq = wq.where(VKWeeklyStats.week_start >= cutoff)
    wq = wq.order_by(VKWeeklyStats.week_start.asc())
    weekly_rows = (await db.execute(wq)).scalars().all()

    if not weekly_rows:
        raise HTTPException(404, "Нет данных ВКонтакте. Сначала соберите аналитику.")

    all_posts: list[dict] = []
    for r in weekly_rows:
        for p in (r.posts or []):
            all_posts.append(p)

    avg_views = sum(p.get("views", 0) or 0 for p in all_posts) / len(all_posts) if all_posts else 0
    avg_er = (
        sum((p.get("likes", 0) + p.get("comments", 0) + p.get("reposts", 0)) / max(p.get("views", 1), 1) * 100
            for p in all_posts) / len(all_posts)
    ) if all_posts else 0

    rated  = _rate_vk_posts_for_analysis(all_posts, avg_views, avg_er)
    mature = [p for p in rated if p["age_days"] >= 7]
    top5   = rated[:5]
    worst5 = list(reversed(mature[-5:])) if mature else []

    week_lines = []
    for r in weekly_rows:
        s = r.stats or {}
        week_lines.append(
            f"  {r.week_start.date()} — {r.week_end.date()}: "
            f"постов={s.get('posts_count',0)}, участников={s.get('members','?')}, "
            f"ср.просм.={s.get('avg_views','?')}, медиана={s.get('median_views','?')}, "
            f"ER подп={s.get('er_subscribers_pct','?')}%, ER просм={s.get('er_views_pct','?')}%, "
            f"лайки={s.get('avg_likes',0):.1f}, репосты={s.get('avg_reposts',0):.1f}, "
            f"вираль={s.get('virality_pct',0):.3f}%"
        )

    def _fmt_post(p: dict, rank: int) -> str:
        age_note = f" [{p['age_label']}]" if p.get("age_label") else ""
        preview  = (p.get("text") or "").replace("\n", " ")[:140] or "(без текста)"
        return (
            f"  #{rank}. id={p.get('id','?')}, рейтинг={p['score']}, "
            f"просм={p.get('views',0)} (расч.~{p['adj_views']}), ER={p['er_pct']:.2f}%{age_note}\n"
            f"     {preview}\n"
            f"     лайки={p.get('likes',0)}, репосты={p.get('reposts',0)}, комм={p.get('comments',0)}"
        )

    top_txt   = "\n".join(_fmt_post(p, i + 1) for i, p in enumerate(top5))   or "  Нет данных"
    worst_txt = "\n".join(_fmt_post(p, i + 1) for i, p in enumerate(worst5)) or "  Нет данных"

    group_name   = weekly_rows[-1].group_name or "ВКонтакте"
    period_label = f"последние {weeks} нед." if weeks > 0 else "всё время"

    prompt = f"""Ты опытный SMM-аналитик. Проведи развёрнутый разбор сообщества ВКонтакте «{group_name}» за {period_label}.

## Недельная статистика (хронологически):
{chr(10).join(week_lines)}

## Топ-5 лучших постов (рейтинг учитывает возраст поста):
{top_txt}

## Топ-5 слабых постов (только зрелые, 7+ дней):
{worst_txt}

Сводка: {len(all_posts)} постов за период, ср. просмотры = {avg_views:.1f}, ср. ER = {avg_er:.2f}%

---

Напиши аналитический отчёт на русском. Структура обязательна:

## 1. Общая оценка сообщества
2-3 предложения — общий вывод о состоянии, тональность честная.

## 2. Динамика метрик
Что растёт, что падает, переломные моменты. Конкретные цифры и недели.

## 3. Анализ контента
Почему топ-посты зашли. Почему слабые провалились. Паттерны.

## 4. Что повторить
Конкретные форматы, темы, подходы — с обоснованием из данных.

## 5. Что исключить
Антипаттерны, форматы с низким ER/просмотрами.

## 6. Рекомендации на следующие 4 недели
Минимум 5 конкретных действий с ожидаемым эффектом.

Пиши честно. Если данных мало — скажи. Не придумывай цифры которых нет в условии."""

    try:
        from anthropic import AsyncAnthropic
        msg = await AsyncAnthropic().messages.create(
            model="claude-sonnet-4-6",
            max_tokens=4000,
            messages=[{"role": "user", "content": prompt}],
        )
        analysis_text = msg.content[0].text.strip()
    except Exception as e:
        raise HTTPException(500, f"Ошибка AI-анализа: {e}")

    def _post_out(p: dict) -> dict:
        return {
            "post_id": p.get("id", 0),
            "published_at": p.get("date", ""),
            "text": p.get("text", ""),
            "views": p.get("views", 0),
            "adj_views": p["adj_views"],
            "reactions": p.get("likes", 0),
            "comments": p.get("comments", 0),
            "reposts": p.get("reposts", 0),
            "er_pct": p["er_pct"],
            "score": p["score"],
            "age_days": p["age_days"],
            "age_label": p.get("age_label"),
            "media_type": "none",
        }

    return {
        "analysis": analysis_text,
        "top_posts": [_post_out(p) for p in top5],
        "worst_posts": [_post_out(p) for p in worst5],
        "period": period_label,
        "channel_name": group_name,
        "total_posts": len(all_posts),
        "avg_views": round(avg_views, 1),
        "avg_er": round(avg_er, 2),
    }


def _rate_posts_for_analysis(posts: list[dict], avg_views: float, avg_er: float) -> list[dict]:
    """Рейтинг постов с поправкой на возраст — свежие посты ещё не набрали полную аудиторию."""
    from datetime import datetime as _dt
    now = _dt.utcnow()
    rated = []
    for p in posts:
        try:
            pub_str = p["published_at"].replace("Z", "").split(".")[0]
            published = _dt.fromisoformat(pub_str)
        except Exception:
            published = now
        age_days = max((now - published).days, 0)

        views   = p.get("views", 0) or 0
        er      = p.get("er_pct", 0) or 0
        viral   = p.get("virality_pct", 0) or 0

        # Поправочный коэффициент: свежие посты не успели набрать просмотры
        if age_days < 1:
            age_factor, age_label = 0.08, "опубликован сегодня — данных слишком мало"
        elif age_days < 3:
            age_factor, age_label = 0.30, f"{age_days} дн. назад — ещё набирает просмотры"
        elif age_days < 7:
            age_factor, age_label = 0.65, f"{age_days} дн. назад — статистика неполная"
        elif age_days < 14:
            age_factor, age_label = 0.90, None
        else:
            age_factor, age_label = 1.0, None

        adj_views  = views / age_factor if age_factor > 0 else views
        views_score = min((adj_views / max(avg_views, 1)) * 50, 100)
        er_score    = min((er / max(avg_er, 0.01)) * 30, 60)
        viral_score = min(viral * 5, 20)
        score = round(views_score + er_score + viral_score, 1)

        rated.append({**p, "score": score, "age_days": age_days,
                      "age_label": age_label, "adj_views": round(adj_views)})

    return sorted(rated, key=lambda x: x["score"], reverse=True)


@router.post("/{business_id}/tg/ai-analysis")
async def tg_ai_analysis(
    business_id: str,
    weeks: int = Query(8, description="Кол-во недель для анализа; 0 = всё время"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Полный AI-разбор канала: динамика, рейтинг постов, рекомендации."""
    await _get_business(business_id, current_user, db)

    from datetime import date, timedelta, datetime as _dt

    # ── Недельная статистика ──────────────────────────────────────────────────
    wq = select(TGWeeklyStats).where(TGWeeklyStats.business_id == business_id)
    if weeks > 0:
        cutoff = date.today() - timedelta(weeks=weeks)
        wq = wq.where(TGWeeklyStats.week_start >= cutoff)
    wq = wq.order_by(TGWeeklyStats.week_start.asc())
    weekly_rows = (await db.execute(wq)).scalars().all()

    if not weekly_rows:
        raise HTTPException(404, "Нет недельных данных. Сначала соберите аналитику.")

    # ── Посты ─────────────────────────────────────────────────────────────────
    pq = select(TelegramPost).where(TelegramPost.business_id == business_id)
    if weeks > 0:
        cutoff_dt = _dt.utcnow() - timedelta(weeks=weeks)
        pq = pq.where(TelegramPost.published_at >= cutoff_dt)
    pq = pq.order_by(TelegramPost.published_at.desc()).limit(500)
    posts_data = [
        {
            "post_id": p.post_id,
            "published_at": p.published_at.isoformat(),
            "text": (p.text or "")[:300],
            "views": p.views or 0,
            "reactions": p.reactions or 0,
            "comments": p.comments or 0,
            "reposts": p.reposts or 0,
            "er_pct": float(p.er_pct or 0),
            "virality_pct": float(p.virality_pct or 0),
            "has_media": p.has_media,
            "media_type": p.media_type or "none",
            "has_question": p.has_question,
        }
        for p in (await db.execute(pq)).scalars().all()
    ]

    avg_views = sum(p["views"] for p in posts_data) / len(posts_data) if posts_data else 0
    avg_er    = sum(p["er_pct"] for p in posts_data) / len(posts_data) if posts_data else 0

    rated  = _rate_posts_for_analysis(posts_data, avg_views, avg_er)
    mature = [p for p in rated if p["age_days"] >= 7]
    top5   = rated[:5]
    worst5 = list(reversed(mature[-5:])) if mature else []

    # ── Строим текст для промпта ──────────────────────────────────────────────
    week_lines = []
    for r in weekly_rows:
        s = r.stats or {}
        week_lines.append(
            f"  {r.week_start.date()} — {r.week_end.date()}: "
            f"постов={s.get('posts_count',0)}, подп.={s.get('subscribers','?')}, "
            f"ср.просм.={s.get('avg_views','?')}, медиана={s.get('median_views','?')}, "
            f"ER просм={s.get('er_reach_pct','?')}%, ER акт={s.get('er_activity_pct','?')}%, "
            f"реакц={s.get('avg_reactions',0):.1f}, репосты={s.get('avg_reposts',0):.1f}, "
            f"вираль={s.get('virality_pct',0):.3f}%"
        )

    def _fmt_post(p: dict, rank: int) -> str:
        age_note = f" [{p['age_label']}]" if p.get("age_label") else ""
        preview  = (p["text"] or "").replace("\n", " ")[:140] or "(без текста)"
        return (
            f"  #{rank}. post_id={p['post_id']}, рейтинг={p['score']}, "
            f"просм={p['views']} (расч.~{p['adj_views']}), ER={p['er_pct']:.2f}%{age_note}\n"
            f"     {preview}\n"
            f"     медиа={p['media_type']}, реакц={p['reactions']}, "
            f"репосты={p['reposts']}, вопрос={'да' if p['has_question'] else 'нет'}"
        )

    top_txt   = "\n".join(_fmt_post(p, i + 1) for i, p in enumerate(top5))   or "  Нет данных"
    worst_txt = "\n".join(_fmt_post(p, i + 1) for i, p in enumerate(worst5)) or "  Нет данных"

    channel_name  = weekly_rows[-1].channel_name or "Telegram-канал"
    period_label  = f"последние {weeks} нед." if weeks > 0 else "всё время"

    prompt = f"""Ты опытный SMM-аналитик. Проведи развёрнутый разбор Telegram-канала «{channel_name}» за {period_label}.

## Недельная статистика (хронологически):
{chr(10).join(week_lines)}

## Топ-5 лучших постов (рейтинг учитывает возраст поста):
{top_txt}

## Топ-5 слабых постов (только зрелые, 7+ дней):
{worst_txt}

Сводка: {len(posts_data)} постов за период, ср. просмотры = {avg_views:.1f}, ср. ER = {avg_er:.2f}%

---

Напиши аналитический отчёт на русском. Структура обязательна:

## 1. Общая оценка канала
2-3 предложения — общий вывод о состоянии, тональность честная.

## 2. Динамика метрик
Что растёт, что падает, переломные моменты. Конкретные цифры и недели.

## 3. Анализ контента
Почему топ-посты зашли (формат, тема, медиа, вопрос?). Почему слабые провалились. Паттерны.

## 4. Что повторить
Конкретные форматы, темы, подходы — с обоснованием из данных.

## 5. Что исключить
Антипаттерны, форматы с низким ER/просмотрами.

## 6. Рекомендации на следующие 4 недели
Минимум 5 конкретных действий с ожидаемым эффектом.

Пиши честно. Если данных мало — скажи. Не придумывай цифры которых нет в условии."""

    try:
        from anthropic import AsyncAnthropic
        msg = await AsyncAnthropic().messages.create(
            model="claude-sonnet-4-6",
            max_tokens=4000,
            messages=[{"role": "user", "content": prompt}],
        )
        analysis_text = msg.content[0].text.strip()
    except Exception as e:
        raise HTTPException(500, f"Ошибка AI-анализа: {e}")

    def _post_out(p: dict) -> dict:
        return {k: p[k] for k in (
            "post_id", "published_at", "text", "views", "adj_views",
            "reactions", "comments", "reposts", "er_pct", "score",
            "age_days", "age_label", "media_type"
        )}

    return {
        "analysis": analysis_text,
        "top_posts": [_post_out(p) for p in top5],
        "worst_posts": [_post_out(p) for p in worst5],
        "period": period_label,
        "channel_name": channel_name,
        "total_posts": len(posts_data),
        "avg_views": round(avg_views, 1),
        "avg_er": round(avg_er, 2),
    }


@router.post("/{business_id}/tg/stories/collect")
async def collect_tg_stories(
    business_id: str,
    deeper: bool = Query(False, description="True — уходим глубже в историю"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_business(business_id, current_user, db)
    conn = await _get_connection(business_id, "telegram", db)
    if not conn or not conn.tg_session_encrypted:
        raise HTTPException(400, "Telegram не авторизован. Войдите через вкладку Аналитика → Telegram.")

    from app.config import settings
    api_id = int(settings.TG_API_ID) if settings.TG_API_ID else 0
    api_hash = settings.TG_API_HASH or ""
    if not api_id:
        raise HTTPException(400, "TG_API_ID / TG_API_HASH не настроены.")

    session_str = decrypt_token(conn.tg_session_encrypted)
    channel = conn.external_page_id or ""
    if not channel:
        raise HTTPException(400, "Идентификатор канала не найден. Переподключите Telegram.")

    offset_id = 0
    if deeper:
        res = await db.execute(
            select(func.min(TelegramStory.story_id)).where(TelegramStory.business_id == business_id)
        )
        min_stored = res.scalar()
        if min_stored:
            offset_id = min_stored - 1

    try:
        stories = await asyncio.to_thread(
            _collect_stories_sync, api_id, api_hash, session_str, channel, offset_id
        )
        count = await _upsert_stories(business_id, stories, db)
        return {"status": "ok", "collected": count, "deeper": deeper}
    except Exception as e:
        raise HTTPException(400, f"Ошибка сбора историй: {str(e)}")
